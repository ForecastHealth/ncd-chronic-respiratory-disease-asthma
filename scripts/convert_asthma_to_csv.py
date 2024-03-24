"""
convert_asthma_to_csv.py

A script that goes through the Asthma.xlsx workbook,
and collects the data in a CSV format, so it can be 
easily accessed by a botech Datafetcher.
"""
import pandas as pd
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from typing import List
WORKBOOK_FILEPATH = "./data/Asthma.xlsx"
AGE_BINS = [
    "0_4",
    "5_9",
    "10_14",
    "15_19",
    "20_24",
    "25_29",
    "30_39",
    "40_49",
    "50_59",
    "60_69",
    "70_79",
    "80_100"
]


def get_records(ws: Worksheet) -> List[dict]:
    records = []
    max_column = ws.max_column

    for row_num, _ in enumerate(
        ws.iter_rows(min_col=1, max_col=max_column),
        start=1
    ):
        if ws.cell(row=row_num, column=1).value == "<State Association>":
            source_label = ws.cell(row=row_num + 2, column=3).value
            for record in get_disability_records(ws, row_num, source_label):
                records.append(record)
            for record in get_state_transition_records(ws, row_num, source_label):
                records.append(record)
            for record in get_prevalence_records(ws, row_num, source_label):
                records.append(record)
        if ws.cell(row=row_num, column=1).value == "<Treatment Association>":
            treatment_label = ws.cell(row=row_num + 2, column=3).value
            for record in get_treatment_transition_records(ws, row_num, treatment_label):
                records.append(record)
            for record in get_disability_records(ws, row_num, source_label):
                records.append(record)
    return records


def get_state_transition_records(
    ws: Worksheet,
    row_num: int,
    label: str,
) -> List[dict]:
    start_row = row_num + 5
    max_column = ws.max_column
    for col_num in range(1, max_column + 1):
        cell_value = ws.cell(row=start_row, column=col_num).value
        if cell_value == '<TransName>':
            target = ws.cell(row=start_row + 2, column=col_num + 2).value
            transition_rates = [ws.cell(
                row=r,
                column=col_num + 1
            ).value for r in range(
                start_row + 2,
                start_row + 19
            )]
            transition_rates = [rate if rate is not None else 0 for rate in transition_rates]
            if sum(transition_rates) == 0:
                continue
            for i in range(len(AGE_BINS)):
                yield {
                        "observation": f"{label} -> {target}",
                        "age": AGE_BINS[i],
                        "value": transition_rates[i]}


def get_prevalence_records(
    ws: Worksheet,
    row_num: int,
    label: str,
) -> List[dict]:
    start_row = row_num + 5
    max_column = ws.max_column
    for col_num in range(1, max_column + 1):
        cell_value = ws.cell(row=start_row, column=col_num).value
        if cell_value == '<Baseline Prevalence>':
            transition_rates = [ws.cell(
                row=r,
                column=col_num
            ).value for r in range(
                start_row + 2,
                start_row + 19
            )]
            transition_rates = [rate if rate is not None else 0 for rate in transition_rates]
            if sum(transition_rates) == 0:
                continue
            for i in range(len(AGE_BINS)):
                yield {
                        "observation": f"Prevalence -> {label}",
                        "age": AGE_BINS[i],
                        "value": transition_rates[i]}


def get_treatment_transition_records(
    ws: Worksheet,
    row_num: int,
    label: str,
) -> List[dict]:
    start_row = row_num + 6
    max_column = ws.max_column
    for col_num in range(1, max_column + 1):
        cell_value = ws.cell(row=start_row, column=col_num).value
        if cell_value == '<Trans ID>':
            target = ws.cell(row=start_row + 2, column=col_num + 1).value
            if target is None or label is None:
                continue
            transition_rates = [ws.cell(
                row=r,
                column=col_num + 2
            ).value for r in range(
                start_row + 2,
                start_row + 19
            )]
            for i in range(len(AGE_BINS)):
                yield {
                        "observation": f"{label} -> {target}",
                        "age": AGE_BINS[i],
                        "value": transition_rates[i] / 100}


def get_disability_records(
    ws: Worksheet,
    row_num: int,
    label: str
):
    if label != "DsFreeSus":
        candidate_label = ws.cell(
            row=row_num + 3,
            column=2
        ).value
        if candidate_label == "DW":
            disability_weight = ws.cell(
                row=row_num + 3,
                column=3
            ).value
            for i in range(len(AGE_BINS)):
                yield {
                        "observation": f"{label} -> Disability",
                        "age": AGE_BINS[i],
                        "value": disability_weight}

    elif label == "DsFreeSus":
        start_row = row_num + 5
        col_num = 3
        transition_rates = [ws.cell(
            row=r,
            column=col_num + 1
        ).value for r in range(
            start_row + 2,
            start_row + 19
        )]
        for i in range(len(AGE_BINS)):
            yield {
                    "observation": f"{label} -> Disability",
                    "age": AGE_BINS[i],
                    "value": transition_rates[i]
            }


def main():
    workbook = openpyxl.load_workbook(WORKBOOK_FILEPATH)
    worksheets = workbook.sheetnames
    records = []
    for worksheet in worksheets:
        if worksheet == "Sheet1":
            continue
        region, sex = worksheet.split("_")
        xl = workbook[worksheet]
        ws_records = get_records(xl)
        for record in ws_records:
            record["region"] = region
            record["sex"] = sex
        records.extend(ws_records)
    df = pd.DataFrame(records)
    df.to_csv("./data/asthma.csv", index=False)


if __name__ == "__main__":
    main()
